from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from evaluation.build_cost_model import DEFAULT_OUTPUT, build_workbook, write_cost_model
from evaluation.cost_model_helpers import ensure_pricing_yaml, load_pricing_raw


EXPECTED_SHEETS = [
    "Legend",
    "Assumptions",
    "Eval Results",
    "Cost Per Doc",
    "Scaling Projection",
    "Dashboard",
]


def test_ensure_pricing_yaml_creates_when_missing(tmp_path: Path):
    path = tmp_path / "pricing.yaml"
    assert not path.exists()
    created = ensure_pricing_yaml(path)
    assert created == path and path.exists()
    raw = load_pricing_raw(path, create_if_missing=False)
    assert "anthropic" in raw["frontier_models"]
    assert raw["local_compute"]["gpu_hourly_rate_usd"] == 0.8


def test_load_pricing_raw_falls_back_on_corrupt_file(tmp_path: Path):
    path = tmp_path / "pricing.yaml"
    path.write_text("not: valid: yaml: [[[", encoding="utf-8")
    raw = load_pricing_raw(path, create_if_missing=False)
    assert raw["frontier_models"]["openai"]["input_per_million"] == 2.5


def test_build_workbook_sheet_structure():
    wb = build_workbook()
    assert wb.sheetnames == EXPECTED_SHEETS

    assumptions = wb["Assumptions"]
    assert assumptions["B9"].value is not None  # GPU hourly rate
    assert assumptions["C5"].value == 3.0 or float(assumptions["C5"].value) > 0

    cost = wb["Cost Per Doc"]
    # First Anthropic classification cost is a formula
    assert str(cost["I4"].value).startswith("=")
    # Local row uses GPU / throughput
    assert "Assumptions!B9" in str(cost["G6"].value)

    scaling = wb["Scaling Projection"]
    assert "Cost Per Doc" in str(scaling["B5"].value)

    dash = wb["Dashboard"]
    assert "Scaling Projection" in str(dash["B5"].value)
    assert dash._charts  # noqa: SLF001 — chart present on dashboard


def test_write_cost_model(tmp_path: Path):
    out = tmp_path / "cost_model.xlsx"
    written = write_cost_model(out)
    assert written == out
    assert out.exists() and out.stat().st_size > 1000

    wb = load_workbook(out)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_default_output_path_under_evaluation():
    assert DEFAULT_OUTPUT.as_posix().endswith("evaluation/cost_model/cost_model.xlsx")
