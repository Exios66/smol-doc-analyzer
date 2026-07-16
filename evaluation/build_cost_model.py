"""
Builds cost_model.xlsx: a template for comparing frontier vs. local model
costs across the three eval tasks, driven by formulas so AmFam's team (or you)
can drop in real pricing/eval numbers and have everything recalculate.

Usage:
    python -m evaluation.build_cost_model
    python -m evaluation.build_cost_model --output evaluation/cost_model/cost_model.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.config import REPO_ROOT

BLUE = Font(name="Arial", color="0000FF", size=10)  # hardcoded inputs
BLACK = Font(name="Arial", color="000000", size=10)  # formulas
BOLD_BLACK = Font(name="Arial", color="000000", size=10, bold=True)
GREEN = Font(name="Arial", color="008000", size=10)  # cross-sheet links
HEADER_FONT = Font(name="Arial", color="FFFFFF", size=11, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_4 = "$#,##0.0000;($#,##0.0000);-"
CURRENCY_2 = "$#,##0.00;($#,##0.00);-"
CURRENCY_0 = "$#,##0;($#,##0);-"
PCT_1 = "0.0%"

DEFAULT_OUTPUT = REPO_ROOT / "evaluation" / "cost_model" / "cost_model.xlsx"
PRICING_YAML = REPO_ROOT / "evaluation" / "pricing.yaml"


def style_header_row(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _load_pricing_defaults() -> dict:
    """Seed Assumptions from evaluation/pricing.yaml when available."""
    defaults = {
        "anthropic_in": 3.00,
        "anthropic_out": 15.00,
        "openai_in": 2.50,
        "openai_out": 10.00,
        # Conservative sustained throughput for the spreadsheet (docs/hour).
        # Distinct from pricing.yaml peak reference_docs_per_hour used by helpers.
        "gpu_hourly_rate": 2.50,
        "throughput": {
            "Classification": 1200,
            "Extraction": 300,
            "Memo Generation": 150,
        },
    }
    try:
        from evaluation.cost_model_helpers import load_pricing_raw

        raw = load_pricing_raw(PRICING_YAML, create_if_missing=True)
        frontier = raw.get("frontier_models") or {}
        if "anthropic" in frontier:
            defaults["anthropic_in"] = float(frontier["anthropic"]["input_per_million"])
            defaults["anthropic_out"] = float(frontier["anthropic"]["output_per_million"])
        if "openai" in frontier:
            defaults["openai_in"] = float(frontier["openai"]["input_per_million"])
            defaults["openai_out"] = float(frontier["openai"]["output_per_million"])
        local = raw.get("local_compute") or {}
        if "gpu_hourly_rate_usd" in local:
            defaults["gpu_hourly_rate"] = float(local["gpu_hourly_rate_usd"])
        # Spreadsheet keeps conservative sustained rates; do not overwrite with
        # peak reference_docs_per_hour (those are 10x higher and inflate savings).
    except Exception as exc:  # noqa: BLE001
        import logging

        logging.getLogger(__name__).warning(
            "Could not seed cost model from %s (%s); using hardcoded defaults",
            PRICING_YAML,
            exc,
        )
    return defaults


def build_workbook() -> Workbook:
    """Construct the multi-sheet cost model workbook (formulas, not cached values)."""
    defaults = _load_pricing_defaults()
    wb = Workbook()

    # ============================================================ Sheet 1: Legend
    ws = wb.active
    ws.title = "Legend"
    ws["A1"] = "AmFam Document AI -- Local vs. Frontier Model Cost Model"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A3"] = "How to use this workbook"
    ws["A3"].font = BOLD_BLACK
    instructions = [
        "1. Fill in blue cells on the 'Assumptions' tab with real pricing and measured throughput.",
        "2. Fill in blue cells on the 'Eval Results' tab with accuracy/F1 numbers from metrics.py output.",
        "3. All other sheets (Cost Per Doc, Scaling Projection, Dashboard) recalculate automatically.",
        "4. Color key: blue text = input you edit. Black = formula, do not overwrite.",
        "   Green = link to another sheet. Yellow fill = key assumption to confirm before presenting.",
        "5. Open in Excel or LibreOffice after editing so cached formula values refresh.",
        "6. Defaults seeded from evaluation/pricing.yaml — confirm yellow cells before presenting.",
        "7. Paste scores from evaluation/metrics.py summary.csv into 'Eval Results'.",
    ]
    for i, line in enumerate(instructions, start=4):
        ws.cell(row=i, column=1, value=line).font = BLACK
    autosize(ws, {"A": 100})

    # ============================================================ Sheet 2: Assumptions
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Assumptions & Inputs"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)

    # --- Frontier pricing ---
    ws["A3"] = "Frontier Model Pricing (USD per 1M tokens)"
    ws["A3"].font = BOLD_BLACK
    ws["A3"].fill = SUBHEAD_FILL
    headers = ["Backend", "Model", "Input $/1M tok", "Output $/1M tok", "Source / date checked"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    pricing_rows = [
        (
            "Anthropic",
            "claude-sonnet-4.5",
            defaults["anthropic_in"],
            defaults["anthropic_out"],
            "anthropic.com/pricing -- CHECK & date",
        ),
        (
            "OpenAI",
            "gpt-4o",
            defaults["openai_in"],
            defaults["openai_out"],
            "openai.com/api/pricing -- CHECK & date",
        ),
    ]
    for r, (backend, model, in_price, out_price, source) in enumerate(pricing_rows, start=5):
        ws.cell(row=r, column=1, value=backend).font = BLACK
        ws.cell(row=r, column=2, value=model).font = BLACK
        c = ws.cell(row=r, column=3, value=in_price)
        c.font = BLUE
        c.fill = ASSUMPTION_FILL
        c.number_format = CURRENCY_2
        c = ws.cell(row=r, column=4, value=out_price)
        c.font = BLUE
        c.fill = ASSUMPTION_FILL
        c.number_format = CURRENCY_2
        ws.cell(row=r, column=5, value=source).font = BLUE

    # --- Local compute ---
    ws["A8"] = "Local Compute Assumptions"
    ws["A8"].font = BOLD_BLACK
    ws["A8"].fill = SUBHEAD_FILL
    ws["A9"] = "GPU hourly rate (USD)"
    ws["A9"].font = BLACK
    c = ws["B9"]
    c.value = defaults["gpu_hourly_rate"]
    c.font = BLUE
    c.fill = ASSUMPTION_FILL
    c.number_format = CURRENCY_2
    ws["C9"] = (
        "Source: seeded from evaluation/pricing.yaml (or CHTC / cloud on-demand) -- CONFIRM"
    )
    ws["C9"].font = BLUE

    # --- Per-task token / throughput estimates ---
    ws["A11"] = "Per-Task Volume Estimates"
    ws["A11"].font = BOLD_BLACK
    ws["A11"].fill = SUBHEAD_FILL
    headers2 = [
        "Task",
        "Avg input tok/doc (frontier)",
        "Avg output tok/doc (frontier)",
        "Local throughput (docs/hour)",
    ]
    for i, h in enumerate(headers2, start=1):
        ws.cell(row=12, column=i, value=h)
    style_header_row(ws, 12, len(headers2))

    task_rows = [
        ("Classification", 800, 20, defaults["throughput"]["Classification"]),
        ("Extraction", 1500, 400, defaults["throughput"]["Extraction"]),
        ("Memo Generation", 2000, 600, defaults["throughput"]["Memo Generation"]),
    ]
    for r, (task, in_tok, out_tok, throughput) in enumerate(task_rows, start=13):
        ws.cell(row=r, column=1, value=task).font = BLACK
        for col, val in zip((2, 3, 4), (in_tok, out_tok, throughput)):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BLUE
            c.fill = ASSUMPTION_FILL

    # --- Volume tiers for scaling projection ---
    ws["A17"] = "Monthly Volume Tiers (docs/month, for scaling projection)"
    ws["A17"].font = BOLD_BLACK
    ws["A17"].fill = SUBHEAD_FILL
    tiers = [1000, 10000, 50000, 100000, 500000]
    for i, t in enumerate(tiers, start=1):
        c = ws.cell(row=18, column=i, value=t)
        c.font = BLUE
        c.fill = ASSUMPTION_FILL
        c.number_format = "#,##0"

    autosize(ws, {"A": 26, "B": 24, "C": 24, "D": 24, "E": 34})

    # ============================================================ Sheet 3: Eval Results
    ws = wb.create_sheet("Eval Results")
    ws["A1"] = "Eval Results (fill in from metrics.py summary.csv output)"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)
    headers3 = ["Task", "Backend", "Accuracy / Macro-F1", "N examples", "Notes"]
    for i, h in enumerate(headers3, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers3))

    eval_rows = [
        ("Classification", "Anthropic", 0.94, 200, "example row -- replace with real run"),
        ("Classification", "OpenAI", 0.93, 200, ""),
        ("Classification", "Local (DeBERTa-v3)", 0.91, 200, ""),
        ("Extraction", "Anthropic", 0.88, 200, ""),
        ("Extraction", "OpenAI", 0.86, 200, ""),
        ("Extraction", "Local (LayoutLMv3)", 0.83, 200, ""),
        ("Memo Generation", "Anthropic", 0.90, 100, "rubric coverage, not accuracy"),
        ("Memo Generation", "OpenAI", 0.88, 100, ""),
        ("Memo Generation", "Local (LoRA)", 0.81, 100, ""),
    ]
    for r, (task, backend, score, n, notes) in enumerate(eval_rows, start=4):
        ws.cell(row=r, column=1, value=task).font = BLACK
        ws.cell(row=r, column=2, value=backend).font = BLACK
        c = ws.cell(row=r, column=3, value=score)
        c.font = BLUE
        c.fill = ASSUMPTION_FILL
        c.number_format = PCT_1
        c = ws.cell(row=r, column=4, value=n)
        c.font = BLUE
        c.fill = ASSUMPTION_FILL
        ws.cell(row=r, column=5, value=notes).font = BLUE

    autosize(ws, {"A": 20, "B": 22, "C": 22, "D": 14, "E": 34})

    # ============================================================ Sheet 4: Cost Per Document
    ws = wb.create_sheet("Cost Per Doc")
    ws["A1"] = "Cost Per Document by Task & Backend"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)
    headers4 = [
        "Task",
        "Backend",
        "Input tok/doc",
        "Output tok/doc",
        "Input $/1M",
        "Output $/1M",
        "GPU $/hr",
        "Docs/hr (local)",
        "Cost / Doc ($)",
    ]
    for i, h in enumerate(headers4, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers4))

    # Row layout mirrors Assumptions!13:15 (task order) x 3 backends each
    task_names = ["Classification", "Extraction", "Memo Generation"]
    backends = [("Anthropic", 5), ("OpenAI", 6), ("Local", None)]  # pricing rows

    row = 4
    for t_idx, task in enumerate(task_names):
        assumption_row = 13 + t_idx  # Assumptions!A13:D15
        for backend_name, pricing_row in backends:
            ws.cell(row=row, column=1, value=task).font = BLACK
            ws.cell(row=row, column=2, value=backend_name).font = BLACK

            in_tok_cell = f"Assumptions!B{assumption_row}"
            out_tok_cell = f"Assumptions!C{assumption_row}"
            throughput_cell = f"Assumptions!D{assumption_row}"

            c = ws.cell(row=row, column=3, value=f"={in_tok_cell}")
            c.font = GREEN
            c = ws.cell(row=row, column=4, value=f"={out_tok_cell}")
            c.font = GREEN

            if backend_name != "Local":
                c = ws.cell(row=row, column=5, value=f"=Assumptions!C{pricing_row}")
                c.font = GREEN
                c.number_format = CURRENCY_2
                c = ws.cell(row=row, column=6, value=f"=Assumptions!D{pricing_row}")
                c.font = GREEN
                c.number_format = CURRENCY_2
                ws.cell(row=row, column=7, value="-").font = BLACK
                ws.cell(row=row, column=8, value="-").font = BLACK
                cost_formula = f"=(C{row}/1000000)*E{row}+(D{row}/1000000)*F{row}"
            else:
                ws.cell(row=row, column=5, value="-").font = BLACK
                ws.cell(row=row, column=6, value="-").font = BLACK
                c = ws.cell(row=row, column=7, value="=Assumptions!B9")
                c.font = GREEN
                c.number_format = CURRENCY_2
                c = ws.cell(row=row, column=8, value=f"={throughput_cell}")
                c.font = GREEN
                cost_formula = f"=IFERROR(G{row}/H{row},0)"

            c = ws.cell(row=row, column=9, value=cost_formula)
            c.font = BOLD_BLACK
            c.number_format = CURRENCY_4
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = BORDER
            row += 1

    autosize(ws, {c: 16 for c in "ABCDEFGHI"})
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["I"].width = 16

    # ============================================================ Sheet 5: Scaling Projection
    ws = wb.create_sheet("Scaling Projection")
    ws["A1"] = "Monthly Cost at Volume (USD)"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)
    ws["A2"] = (
        "Uses average cost/doc across all three tasks per backend, x monthly volume tier."
    )
    ws["A2"].font = BLACK

    headers5 = (
        ["Volume tier (docs/month)"]
        + [f"{n} avg $/doc" for n in ["Anthropic", "OpenAI", "Local"]]
        + [
            "Anthropic $/mo",
            "OpenAI $/mo",
            "Local $/mo",
            "Savings: Local vs Anthropic",
            "Savings: Local vs OpenAI",
        ]
    )
    for i, h in enumerate(headers5, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers5))

    # Cost Per Doc layout: rows 4-6 = Classification(Anthropic,OpenAI,Local),
    # 7-9 Extraction, 10-12 Memo
    avg_formulas = {
        "Anthropic": "=AVERAGE('Cost Per Doc'!I4,'Cost Per Doc'!I7,'Cost Per Doc'!I10)",
        "OpenAI": "=AVERAGE('Cost Per Doc'!I5,'Cost Per Doc'!I8,'Cost Per Doc'!I11)",
        "Local": "=AVERAGE('Cost Per Doc'!I6,'Cost Per Doc'!I9,'Cost Per Doc'!I12)",
    }

    for r, _ in enumerate(range(5), start=5):
        ws.cell(row=r, column=1, value=f"=Assumptions!{get_column_letter(r - 4)}18")
        ws.cell(row=r, column=1).font = GREEN
        ws.cell(row=r, column=1).number_format = "#,##0"

        c = ws.cell(row=r, column=2, value=avg_formulas["Anthropic"])
        c.font = GREEN
        c.number_format = CURRENCY_4
        c = ws.cell(row=r, column=3, value=avg_formulas["OpenAI"])
        c.font = GREEN
        c.number_format = CURRENCY_4
        c = ws.cell(row=r, column=4, value=avg_formulas["Local"])
        c.font = GREEN
        c.number_format = CURRENCY_4

        c = ws.cell(row=r, column=5, value=f"=A{r}*B{r}")
        c.font = BLACK
        c.number_format = CURRENCY_0
        c = ws.cell(row=r, column=6, value=f"=A{r}*C{r}")
        c.font = BLACK
        c.number_format = CURRENCY_0
        c = ws.cell(row=r, column=7, value=f"=A{r}*D{r}")
        c.font = BOLD_BLACK
        c.number_format = CURRENCY_0

        c = ws.cell(row=r, column=8, value=f"=IFERROR((E{r}-G{r})/E{r},0)")
        c.font = BLACK
        c.number_format = PCT_1
        c = ws.cell(row=r, column=9, value=f"=IFERROR((F{r}-G{r})/F{r},0)")
        c.font = BLACK
        c.number_format = PCT_1

        for col in range(1, 10):
            ws.cell(row=r, column=col).border = BORDER

    autosize(ws, {c: 15 for c in "ABCDEFGHI"})
    ws.column_dimensions["A"].width = 20

    # ============================================================ Sheet 6: Dashboard
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Summary Dashboard"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A3"] = "Headline: cost savings running local models at 50,000 docs/month"
    ws["A3"].font = BOLD_BLACK

    ws["A5"] = "Anthropic monthly cost"
    ws["A5"].font = BLACK
    ws["B5"] = "='Scaling Projection'!E7"
    ws["B5"].font = GREEN
    ws["B5"].number_format = CURRENCY_0
    ws["A6"] = "OpenAI monthly cost"
    ws["A6"].font = BLACK
    ws["B6"] = "='Scaling Projection'!F7"
    ws["B6"].font = GREEN
    ws["B6"].number_format = CURRENCY_0
    ws["A7"] = "Local monthly cost"
    ws["A7"].font = BLACK
    ws["B7"] = "='Scaling Projection'!G7"
    ws["B7"].font = GREEN
    ws["B7"].number_format = CURRENCY_0
    ws["A8"] = "Savings vs. Anthropic"
    ws["A8"].font = BOLD_BLACK
    ws["B8"] = "='Scaling Projection'!H7"
    ws["B8"].font = GREEN
    ws["B8"].number_format = PCT_1
    ws["A9"] = "Savings vs. OpenAI"
    ws["A9"].font = BOLD_BLACK
    ws["B9"] = "='Scaling Projection'!I7"
    ws["B9"].font = GREEN
    ws["B9"].number_format = PCT_1

    for r in range(5, 10):
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER

    # Scaling chart: cost vs volume tier, 3 series
    chart = LineChart()
    chart.title = "Monthly Cost by Volume Tier"
    chart.style = 10
    chart.x_axis.title = "Docs / month"
    chart.y_axis.title = "Monthly cost (USD)"
    sp = wb["Scaling Projection"]
    cats = Reference(sp, min_col=1, min_row=5, max_row=9)
    for col in (5, 6, 7):
        data = Reference(sp, min_col=col, min_row=4, max_row=9)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "D5")

    autosize(ws, {"A": 26, "B": 18})
    return wb


def write_cost_model(output_path: Path | None = None) -> Path:
    """Build and save the workbook; return the path written."""
    output_path = Path(output_path) if output_path else DEFAULT_OUTPUT
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(output_path)
    return output_path


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Build the frontier vs. local cost_model.xlsx spreadsheet template"
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args(argv)
    path = write_cost_model(args.output)
    print(f"saved {path}")


if __name__ == "__main__":
    main()
